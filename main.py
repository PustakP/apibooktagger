import openpyxl
import requests

# Function to read book names from a column in an Excel spreadsheet.
def read_book_names(filename, sheet_name, column_name):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheet_name]
    book_names = []

    for cell in sheet[column_name]:
        book_names.append(cell.value)

    return book_names


# Function to fetch book details from an API
def fetch_book_details(book_name):
    base_url = 'https://www.googleapis.com/books/v1/volumes'
    params = {'q': book_name}

    response = requests.get(base_url, params=params)

    if response.status_code == 200:
        book_data = response.json().get('items', [])
        return book_data[0] if book_data else None
    else:
        return None

# Function to update book details in respective columns in an Excel spreadsheet
def update_book_details(filename, sheet_name, book_names):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheet_name]

    for i in range(len(book_names)):
        book_name = book_names[i]

        # Fetch book details from the API
        book_data = fetch_book_details(book_name)

        if book_data is not None:
            volume_info = book_data.get('volumeInfo', {})
            author_name = volume_info.get('authors', [''])[0]
            tag = volume_info.get('categories', [''])[0]
            summary = volume_info.get('description', '')
            page_count = volume_info.get('pageCount', '')

            sheet.cell(row=i+1, column=3).value = author_name
            sheet.cell(row=i+1, column=4).value = tag
            sheet.cell(row=i+1, column=5).value = summary
            sheet.cell(row=i+1, column=6).value = page_count
            print(i, " is done :p")

    workbook.save(filename)
    workbook.close()

# Example usage
filename = "book.xlsx"
sheet_name = "Sheet1"
column_name = "A"

# Read book names from column A
book_names = read_book_names(filename, sheet_name, column_name)

# Update book details in respective columns
update_book_details(filename, sheet_name, book_names)

